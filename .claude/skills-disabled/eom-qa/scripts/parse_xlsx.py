#!/usr/bin/env python3
"""
Parse EOM reviews XLSX attachment to extract review counts, ratings,
star buckets, and per-address groupings.

Usage:
    python3 parse_xlsx.py <path_to_xlsx>

Output: JSON with extracted values
"""

import json
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_reviews_xlsx(filepath: str) -> dict:
    """Parse the reviews XLSX and return structured data."""
    wb = openpyxl.load_workbook(filepath, read_only=True)

    all_ratings = []
    address_reviews = defaultdict(list)  # address -> list of ratings
    sheet_counts = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Find column indices from header row
        headers = [str(h).strip() if h else '' for h in rows[0]]
        header_lower = [h.lower() for h in headers]

        rating_col = None
        address_col = None

        for i, h in enumerate(header_lower):
            if h == 'rating':
                rating_col = i
            elif h == 'address':
                address_col = i

        if rating_col is None:
            continue

        data_rows = 0
        for row in rows[1:]:
            if row[rating_col] is None:
                continue

            rating = float(row[rating_col])
            all_ratings.append(rating)
            data_rows += 1

            if address_col is not None and row[address_col]:
                addr = str(row[address_col]).strip()
                address_reviews[addr].append(rating)

        sheet_counts[sheet_name] = data_rows

    # Compute stats
    total = len(all_ratings)
    avg_rating = sum(all_ratings) / total if total > 0 else 0
    one_two_star = sum(1 for r in all_ratings if r <= 2)
    three_five_star = sum(1 for r in all_ratings if r >= 3)

    # Star distribution
    star_dist = {}
    for star in range(1, 6):
        star_dist[str(star)] = sum(1 for r in all_ratings if r == star)

    # Per-address stats
    per_address = {}
    for addr, ratings in sorted(address_reviews.items()):
        per_address[addr] = {
            'count': len(ratings),
            'avg': round(sum(ratings) / len(ratings), 2),
            'neg': sum(1 for r in ratings if r <= 2),
            'pos': sum(1 for r in ratings if r >= 3)
        }

    # Verify Positive/Negative sheet assumption
    warnings = []
    if 'Positive' in sheet_counts and 'Negative' in sheet_counts:
        if sheet_counts.get('Positive', 0) != three_five_star:
            warnings.append(
                f"Positive sheet has {sheet_counts['Positive']} rows "
                f"but found {three_five_star} reviews with rating >= 3"
            )
        if sheet_counts.get('Negative', 0) != one_two_star:
            warnings.append(
                f"Negative sheet has {sheet_counts['Negative']} rows "
                f"but found {one_two_star} reviews with rating <= 2"
            )

    wb.close()

    return {
        'total_reviews': total,
        'avg_rating': round(avg_rating, 2),
        'one_two_star': one_two_star,
        'three_five_star': three_five_star,
        'star_distribution': star_dist,
        'sheet_counts': sheet_counts,
        'per_address': per_address,
        'location_count': len(per_address),
        'warnings': warnings
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 parse_xlsx.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)

    result = parse_reviews_xlsx(sys.argv[1])
    print(json.dumps(result, indent=2))
